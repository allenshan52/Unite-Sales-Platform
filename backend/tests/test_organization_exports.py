"""单位 Excel 导出测试：验证审核所需列、筛选结果行和工作簿交付格式。"""

from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.models import CustomerStatus, EvidenceKind, GeocodeStatus, Organization, OrganizationEvidence, OrganizationSite, OrganizationType, ReviewStatus
from app.services.organization_exports import EXPORT_HEADERS, build_organization_export_workbook


def test_export_workbook_keeps_candidate_review_fields() -> None:
    """导出的工作簿必须带地址、审核状态与来源证据，供不登录数据库的人工复核使用。"""

    timestamp = datetime(2026, 8, 6, tzinfo=UTC)
    organization = Organization(
        name="示例理工大学",
        normalized_name="示例理工大学",
        organization_type=OrganizationType.university,
        customer_status=CustomerStatus.potential,
        review_status=ReviewStatus.pending,
        inclusion_reason="材料科学与工程学院",
        website="https://example.test",
        created_at=timestamp,
        updated_at=timestamp,
    )
    organization.sites = [
        OrganizationSite(
            address="示例省示例市示例区学府路 1 号",
            province="示例省",
            city="示例市",
            district="示例区",
            geocode_status=GeocodeStatus.pending,
            is_primary=True,
        )
    ]
    organization.evidences = [
        OrganizationEvidence(
            evidence_kind=EvidenceKind.department,
            title="材料科学与工程学院",
            source_url="https://example.test/materials",
        )
    ]

    workbook = load_workbook(BytesIO(build_organization_export_workbook([organization])))
    sheet = workbook["单位候选"]

    assert tuple(cell.value for cell in sheet[1]) == EXPORT_HEADERS
    assert sheet["A2"].value == "示例理工大学"
    assert sheet["I2"].value == "示例省示例市示例区学府路 1 号"
    assert sheet["J2"].value == "待编码"
    assert sheet["R2"].value == "材料科学与工程学院"
    assert sheet["S2"].value == "https://example.test/materials"
    assert sheet.auto_filter.ref == "A1:U2"


def test_export_workbook_neutralizes_formula_like_text() -> None:
    """管理员和外部来源文本不能在 Excel 打开时被解释为公式。"""

    timestamp = datetime(2026, 8, 6, tzinfo=UTC)
    organization = Organization(
        name="=HYPERLINK(\"https://example.test\")",
        normalized_name="formula-test",
        organization_type=OrganizationType.university,
        customer_status=CustomerStatus.potential,
        review_status=ReviewStatus.pending,
        is_sports_exception=False,
        created_at=timestamp,
        updated_at=timestamp,
        sites=[],
        evidences=[],
    )

    workbook = load_workbook(BytesIO(build_organization_export_workbook([organization])), data_only=False)
    cell = workbook["单位候选"]["A2"]

    assert cell.data_type == "s"
    assert cell.value == "'=HYPERLINK(\"https://example.test\")"
