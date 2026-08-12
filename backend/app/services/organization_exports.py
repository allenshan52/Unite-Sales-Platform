"""单位导出服务：以现有 OpenPyXL 依赖生成可审核的 Excel，不改变数据库记录。"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Organization

EXPORT_HEADERS = (
    "单位名称", "单位类型", "客户状态", "审核状态", "体育例外", "省", "市", "区", "地址", "地理编码状态",
    "经度", "纬度", "行业", "所属集团", "纳入理由", "官网", "证据类型", "证据标题", "证据链接", "录入时间", "最后更新时间",
)
EXPORT_WIDTHS = (26, 12, 14, 12, 11, 12, 12, 12, 34, 14, 13, 13, 18, 18, 42, 30, 16, 32, 42, 20, 20)


def _excel_datetime(value: datetime | None) -> datetime | None:
    """将 PostgreSQL 的带时区时间转为 Excel 可写的 UTC 本地时间，避免导出在保存阶段失败。"""

    if value is None or value.tzinfo is None:
        return value
    return value.astimezone(UTC).replace(tzinfo=None)


def _excel_safe_value(value: object) -> object:
    """把可能被 Excel 当作公式的外部文本强制保存为文字，防止打开导出文件时执行公式。"""

    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{value}"
    return value


def build_organization_export_workbook(organizations: Sequence[Organization]) -> bytes:
    """将按后台筛选条件读取的单位写为工作表，保留地址和来源追溯字段供人工核验。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "单位候选"
    sheet.sheet_view.showGridLines = False
    sheet.append(list(EXPORT_HEADERS))
    for cell in sheet["1:1"]:
        cell.fill = PatternFill("solid", fgColor="1F6B5B")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"

    for organization in organizations:
        site = next((item for item in organization.sites if item.is_primary), organization.sites[0] if organization.sites else None)
        evidences = organization.evidences
        row = [
            organization.name, organization.organization_type.value, organization.customer_status.value,
            organization.review_status.value, "是" if organization.is_sports_exception else "否",
            site.province if site else None, site.city if site else None, site.district if site else None,
            site.address if site else None, site.geocode_status.value if site else None,
            site.longitude if site else None, site.latitude if site else None, organization.industry,
            organization.parent_group, organization.inclusion_reason, organization.website,
            "\n".join(evidence.evidence_kind.value for evidence in evidences),
            "\n".join(evidence.title for evidence in evidences),
            "\n".join(evidence.source_url for evidence in evidences),
            _excel_datetime(organization.created_at), _excel_datetime(organization.updated_at),
        ]
        sheet.append([_excel_safe_value(value) for value in row])

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_HEADERS))}{max(sheet.max_row, 1)}"
    for index, width in enumerate(EXPORT_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 20).number_format = "yyyy-mm-dd hh:mm"
        sheet.cell(row_index, 21).number_format = "yyyy-mm-dd hh:mm"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
