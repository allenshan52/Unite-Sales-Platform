"""数据洞察 Excel 导出：把同一聚合响应写成可复盘的多工作表报表。"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.insights_schemas import InsightsOverviewRead


def _safe(value: object) -> object:
    """阻止单位名称等数据库文本被 Excel 解释为公式。"""

    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{value}"
    return value


def _style_sheet(sheet, widths: tuple[int, ...]) -> None:
    """为业务工作表统一表头、冻结、筛选和中文阅读列宽。"""

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="EF5A2A")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_insights_workbook(overview: InsightsOverviewRead) -> bytes:
    """把当前筛选范围的 KPI、区域、趋势、成交单位和商机阶段导出为 XLSX。"""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "经营概览"
    summary.append(["统计范围", "年份", "期间", "实际销售额（元）", "成交项目数", "平均成交额（元）", "当前商机储备（元）", "有效商机数", "有数据区域数", "聚合时间"])
    summary.append([
        overview.scope.name, overview.year, overview.period.value, overview.kpis.sales_amount,
        overview.kpis.project_count, overview.kpis.average_deal_amount, overview.kpis.pipeline_amount,
        overview.kpis.pipeline_count, overview.kpis.active_region_count, overview.aggregated_at.replace(tzinfo=None),
    ])
    _style_sheet(summary, (18, 10, 12, 20, 14, 20, 22, 14, 14, 22))

    regions = workbook.create_sheet("区域贡献")
    regions.append(["排名", "区域", "省", "市", "实际销售额（元）", "成交项目数", "商机储备（元）", "有效商机数", "指标贡献占比（%）", "同比（%）", "环比（%）"])
    for item in overview.regions:
        regions.append([item.rank, _safe(item.name), _safe(item.province), _safe(item.city or ""), item.sales_amount, item.project_count, item.pipeline_amount, item.pipeline_count, item.contribution_percent, item.yoy_percent, item.qoq_percent])
    _style_sheet(regions, (10, 18, 15, 15, 20, 14, 20, 14, 19, 13, 13))

    trend = workbook.create_sheet("月度趋势")
    trend.append(["月份", f"{overview.year}年实际销售额（元）", f"{overview.year - 1}年实际销售额（元）"])
    for item in overview.trend:
        trend.append([item.month, item.current_amount, item.previous_amount])
    _style_sheet(trend, (10, 24, 24))

    customers = workbook.create_sheet("成交单位")
    customers.append(["排名", "成交单位", "省", "市", "成交金额（元）", "项目数", "最近签约日期"])
    for item in overview.top_customers:
        customers.append([item.rank, _safe(item.name), _safe(item.province), _safe(item.city), item.sales_amount, item.project_count, item.latest_signed_at])
    _style_sheet(customers, (10, 32, 15, 15, 20, 12, 18))

    stages = workbook.create_sheet("商机阶段")
    stages.append(["阶段", "有效商机数", "预计金额（元）", "金额占比（%）"])
    for item in overview.stages:
        stages.append([item.stage, item.opportunity_count, item.amount, item.percent])
    _style_sheet(stages, (18, 16, 20, 18))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
